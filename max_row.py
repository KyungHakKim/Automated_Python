import openpyxl as op 

wb = op.load_workbook(r"max_row.xlsx")
ws = wb.active

#cell에 excel 수식 넣기
ws["E11"].value = "=SUM(C:C)"

col_max = ws.max_column
row_max = ws.max_row

print("row_max = " + str(row_max))
for row in range(2,row_max+1):
    print("row : " + str(row))
    ws["E" + str(row)].value = "=C"+str(row)+"*D"+ str(row)

wb.save(r"row_result.xlsx")

wb2 = op.load_workbook(r"row_result.xlsx", data_only=True)
ws2 = wb2.active

data = []

for row in ws2.rows:
    data.append(row[4].value)

print(data)    


