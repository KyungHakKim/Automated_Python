import openpyxl as op
import win32com.client

def wirteFunc():
    print("Start writeFunc()")
    wb = op.load_workbook(r"max_row.xlsx")
    ws= wb.active
    
    row_max = ws.max_row
    
    for row in range(2, row_max + 1):
        ws["E" + str(row)].value = "=C"+str(row)+"*D"+ str(row)
    
    wb.save(r"max_row_result.xlsx")        

def load_data():
    print("Start load_data()")
    excel = win32com.client.Dispatch("Excel.Application")
    temp_wb = excel.Workbooks.open(r"D:\Python\Automated_Python\max_row_result.xlsx")
    temp_wb.Save()
  
    excel.quit() 
    
    wb = op.load_workbook(r"max_row_result.xlsx", data_only=True)
    ws = wb.active
    
    data = []
    
    for row in ws.rows:
        data.append(row[4].value)
    
    print(data)        

if __name__ == "__main__":    
    print("Start~~~~~~~~~~~~~")
    wirteFunc()
    load_data()
    